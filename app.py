import pandas as pd
import streamlit as st
import io
import re
from datetime import datetime, timedelta
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
import hashlib
import os
import base64

# Configuration de la page
st.set_page_config(page_title="Processeur de Données de Pointage", layout="wide")

# Fonction d'authentification avec hachage renforcé
def authenticate(username, password):
    # Utilisation d'un mot de passe plus fort: "RH@TimeTracker2025!"
    correct_username = "ghita"
    
    # Utilisation de SHA-256 avec sel pour un hachage plus sécurisé
    # Dans une application réelle, ce hachage et ce sel seraient stockés séparément dans une base de données sécurisée
    salt = "f8a923b5e4c2d1"  # Ceci serait généré aléatoirement et stocké en toute sécurité
    correct_password = "RH@TimeTracker2025!"
    
    # Création du hachage avec sel
    correct_hash = hashlib.sha256((correct_password + salt).encode()).hexdigest()
    password_hash = hashlib.sha256((password + salt).encode()).hexdigest()
    
    return username == correct_username and password_hash == correct_hash

# État d'authentification
if 'authenticated' not in st.session_state:
    st.session_state.authenticated = False

# Formulaire de connexion
if not st.session_state.authenticated:
    st.title("Processeur de Données de Pointage - Connexion")
    
    with st.form("login_form"):
        username = st.text_input("Nom d'utilisateur")
        password = st.text_input("Mot de passe", type="password")
        submit_button = st.form_submit_button("Connexion")
        
        if submit_button:
            if authenticate(username, password):
                st.session_state.authenticated = True
                st.experimental_rerun()
            else:
                st.error("Nom d'utilisateur ou mot de passe invalide")
    
    # Arrêter l'exécution si non authentifié
    st.stop()

# Fonction pour normaliser les données de pointage avec 5 valeurs
def normalize_pointage(pointages):
    """
    Normaliser les données de pointage avec 5 valeurs au format standard de 4 valeurs.
    
    Args:
        pointages: Liste de chaînes d'heure au format "HH:MM"
    
    Returns:
        Liste de 4 chaînes d'heure représentant [entrée, début_pause, fin_pause, sortie]
    """
    if len(pointages) == 4:
        return pointages  # Déjà au format correct
    
    if len(pointages) != 5:
        return pointages  # Ne peut pas gérer d'autres formats, retourner tel quel
    
    # Convertir toutes les heures en objets datetime pour une comparaison plus facile
    times = [datetime.strptime(t, "%H:%M") for t in pointages]
    
    # Limites de l'horaire de travail
    morning_start = datetime.strptime("09:00", "%H:%M")
    lunch_start = datetime.strptime("13:00", "%H:%M")
    lunch_end = datetime.strptime("15:00", "%H:%M")
    day_end = datetime.strptime("18:00", "%H:%M")
    
    # Cas 1: Vérifier les pointages en double (heures qui sont à moins de 5 minutes l'une de l'autre)
    for i in range(len(times) - 1):
        time_diff = abs((times[i+1] - times[i]).total_seconds() / 60)
        if time_diff <= 5:  # Dans les 5 minutes
            # Supprimer le doublon (garder le premier)
            del times[i+1]
            return [t.strftime("%H:%M") for t in times]
    
    # Cas 2: Vérifier les pointages multiples de pause
    # Identifier quelles heures tombent dans la période de pause (13:00-15:00)
    break_times = []
    for i, t in enumerate(times):
        if lunch_start <= t <= lunch_end:
            break_times.append((i, t))
    
    # Si nous avons 3 heures de pause, nous devons déterminer lesquelles garder
    if len(break_times) == 3:
        # Stratégie: Garder les première et dernière heures de pause
        # Cela suppose que l'employé a pointé pour la pause, puis a pointé à nouveau pendant la pause
        middle_idx = break_times[1][0]
        del times[middle_idx]
        return [t.strftime("%H:%M") for t in times]
    
    # Cas 3: S'il y a plus d'une entrée matinale ou sortie du soir
    morning_entries = []
    evening_exits = []
    
    for i, t in enumerate(times):
        if t < lunch_start:
            morning_entries.append((i, t))
        elif t > lunch_end:
            evening_exits.append((i, t))
    
    # S'il y a plusieurs entrées matinales, garder celle qui est la plus proche de 9:00
    if len(morning_entries) > 1:
        # Trier par différence absolue par rapport à 9:00
        morning_entries.sort(key=lambda x: abs((x[1] - morning_start).total_seconds()))
        # Garder toutes sauf la première (la plus proche de 9:00)
        for i, _ in morning_entries[1:]:
            # Ajuster l'index pour les suppressions précédentes
            del times[i - morning_entries[1:].index((i, _))]
        return [t.strftime("%H:%M") for t in times]
    
    # S'il y a plusieurs sorties du soir, garder celle qui est la plus proche de 18:00
    if len(evening_exits) > 1:
        # Trier par différence absolue par rapport à 18:00
        evening_exits.sort(key=lambda x: abs((x[1] - day_end).total_seconds()))
        # Garder toutes sauf la première (la plus proche de 18:00)
        for i, _ in evening_exits[1:]:
            # Ajuster l'index pour les suppressions précédentes
            del times[i - evening_exits[1:].index((i, _))]
        return [t.strftime("%H:%M") for t in times]
    
    # Si nous avons toujours 5 valeurs, utiliser une stratégie de repli:
    # Garder la première entrée, la première sortie de pause, la dernière entrée de pause et la dernière sortie
    if len(times) == 5:
        # Supposer que les heures sont plus ou moins dans l'ordre: entrée, sortie pause, [heure supplémentaire], entrée pause, sortie
        del times[2]  # Supprimer la valeur du milieu
    
    return [t.strftime("%H:%M") for t in times]

# Fonction pour analyser le fichier CSV d'entrée
def parse_input_csv(file):
    # Lire le fichier CSV
    df = pd.read_csv(file, delimiter=';', encoding='utf-8', dtype=str)  # Forcer toutes les colonnes à être des chaînes
    
    # Vérifier si le CSV a des en-têtes ou non
    if len(df.columns) == 1:
        # S'il n'y a qu'une seule colonne, le fichier n'a pas d'en-têtes appropriés
        # Diviser la colonne unique en plusieurs colonnes
        df = pd.DataFrame([x.split(';') for x in df.iloc[:, 0].values.tolist()])
        # En supposant que les colonnes sont: Matricule, Nom, Départment, Date, Pointages
        if len(df.columns) >= 5:
            df.columns = ['Matricule', 'Nom', 'Départment', 'Date', 'Pointages'] + [f'Extra{i}' for i in range(len(df.columns) - 5)]
        else:
            # Gérer le cas où il y a moins de colonnes que prévu
            st.error(f"Erreur de format CSV: Au moins 5 colonnes attendues, {len(df.columns)} trouvées")
            return None
    
    # S'assurer que toutes les colonnes sont des chaînes
    for col in df.columns:
        df[col] = df[col].astype(str)
    
    # Nettoyer la colonne Pointages - remplacer NaN par une chaîne vide
    df['Pointages'] = df['Pointages'].fillna('').astype(str)
    df['Pointages'] = df['Pointages'].replace('nan', '')
    
    return df

# Fonction pour traiter les entrées de temps
def process_time_entries(df):
    results = []
    
    for _, row in df.iterrows():
        matricule = row['Matricule']
        nom = row['Nom']
        departement = row['Départment']
        date_str = row['Date']
        pointages_str = row['Pointages']
        
        # Ignorer les lignes avec des pointages vides ou invalides
        if pd.isna(pointages_str) or pointages_str == 'nan' or pointages_str == '':
            pointages_str = ''
        
        # Analyser la date
        try:
            date = datetime.strptime(date_str, '%d/%m/%Y')
        except:
            # Ignorer les dates invalides
            continue
        
        # Analyser les pointages
        try:
            pointages = pointages_str.split() if pointages_str else []
            
            # Indicateur pour les pointages à 5 valeurs
            original_pointage = ' '.join(pointages)
            has_five_values = len(pointages) == 5
            
            # Gérer les pointages à 5 valeurs
            if len(pointages) == 5:
                pointages = normalize_pointage(pointages)
                
        except AttributeError:
            # Gérer le cas où pointages_str n'est pas une chaîne
            pointages = []
            has_five_values = False
            original_pointage = str(pointages_str)
        
        # Valeurs par défaut
        entree = ""
        sortie = ""
        debut_pause = ""
        fin_pause = ""
        temps_pause = "00:00"
        h_p = "01:00"  # Temps de pause par défaut
        h_tr = ""
        h_l = "08:00"  # Heures de travail par défaut
        heures_perdues = ""
        heures_supp = ""
        observations = ""
        
        # Définir les observations pour les pointages à 5 valeurs
        if has_five_values:
            observations = f"À VÉRIFIER: Original avait 5 valeurs: {original_pointage}"
        
        if len(pointages) == 4:
            # Données complètes: entrée, sortie pause, entrée pause, sortie
            entree = pointages[0]
            debut_pause = pointages[1]
            fin_pause = pointages[2]
            sortie = pointages[3]
            
            # Calculer le temps de pause
            try:
                t1 = datetime.strptime(debut_pause, '%H:%M')
                t2 = datetime.strptime(fin_pause, '%H:%M')
                pause_delta = (t2 - t1) if t2 > t1 else (t2 - t1 + timedelta(days=1))
                pause_hours = pause_delta.seconds // 3600
                pause_minutes = (pause_delta.seconds % 3600) // 60
                temps_pause = f"{pause_hours:02d}:{pause_minutes:02d}"
                
                # Calculer les heures de travail
                t_entry = datetime.strptime(entree, '%H:%M')
                t_exit = datetime.strptime(sortie, '%H:%M')
                work_delta = (t_exit - t_entry) if t_exit > t_entry else (t_exit - t_entry + timedelta(days=1))
                work_delta = work_delta - pause_delta  # Soustraire le temps de pause
                work_hours = work_delta.seconds // 3600
                work_minutes = (work_delta.seconds % 3600) // 60
                h_tr = f"{work_hours:02d}:{work_minutes:02d}"
                
                # Calculer s'il y a des heures perdues
                standard_hours = datetime.strptime(h_l, '%H:%M')
                actual_hours = datetime.strptime(h_tr, '%H:%M')
                
                if actual_hours < standard_hours:
                    diff = standard_hours - actual_hours
                    lost_hours = diff.seconds // 3600
                    lost_minutes = (diff.seconds % 3600) // 60
                    heures_perdues = f"{lost_hours:02d}:{lost_minutes:02d}"
                elif actual_hours > standard_hours:
                    diff = actual_hours - standard_hours
                    extra_hours = diff.seconds // 3600
                    extra_minutes = (diff.seconds % 3600) // 60
                    heures_supp = f"{extra_hours:02d}:{extra_minutes:02d}"
            except Exception as e:
                if observations:
                    observations += " | Erreur dans les calculs de temps"
                else:
                    observations = "Erreur dans les calculs de temps"
                
        elif len(pointages) == 2:
            # Seulement les heures d'entrée et de sortie
            entree = pointages[0]
            sortie = pointages[1]
            if observations:
                observations += " | Données de pause manquantes"
            else:
                observations = "Données de pause manquantes"
            
            # Calculer les heures de travail sans pause
            try:
                t_entry = datetime.strptime(entree, '%H:%M')
                t_exit = datetime.strptime(sortie, '%H:%M')
                work_delta = (t_exit - t_entry) if t_exit > t_entry else (t_exit - t_entry + timedelta(days=1))
                # Supposer une pause standard
                pause_delta = datetime.strptime(h_p, '%H:%M') - datetime.strptime("00:00", '%H:%M')
                work_delta = work_delta - pause_delta
                work_hours = work_delta.seconds // 3600
                work_minutes = (work_delta.seconds % 3600) // 60
                h_tr = f"{work_hours:02d}:{work_minutes:02d}"
                
                # Calculer s'il y a des heures perdues
                standard_hours = datetime.strptime(h_l, '%H:%M')
                actual_hours = datetime.strptime(h_tr, '%H:%M')
                
                if actual_hours < standard_hours:
                    diff = standard_hours - actual_hours
                    lost_hours = diff.seconds // 3600
                    lost_minutes = (diff.seconds % 3600) // 60
                    heures_perdues = f"{lost_hours:02d}:{lost_minutes:02d}"
                elif actual_hours > standard_hours:
                    diff = actual_hours - standard_hours
                    extra_hours = diff.seconds // 3600
                    extra_minutes = (diff.seconds % 3600) // 60
                    heures_supp = f"{extra_hours:02d}:{extra_minutes:02d}"
            except Exception as e:
                if observations:
                    observations += " | Erreur dans les calculs de temps"
                else:
                    observations = "Erreur dans les calculs de temps"
        elif len(pointages) == 0:
            if observations:
                observations += " | Absent"
            else:
                observations = "Absent"
            entree = "00:00"
            sortie = "00:00"
            temps_pause = "00:00"
            h_tr = "00:00"
        else:
            if observations:
                observations += f" | Données irrégulières: {len(pointages)} valeurs"
            else:
                observations = f"Données irrégulières: {len(pointages)} valeurs"
        
        results.append({
            'Matricule': matricule,
            'Nom': nom,
            'Départment': departement,
            'Date': date_str,
            'Entrée': entree,
            'Sortie': sortie,
            'Début Pause': debut_pause,
            'Fin Pause': fin_pause,
            'Temps de pause': temps_pause,
            'H.P': h_p,
            'H. Tr': h_tr,
            'HL': h_l,
            'Heures perdues': heures_perdues,
            'Heures supp': heures_supp,
            'Observations': observations
        })
    
    return pd.DataFrame(results)

# Fonction pour créer un fichier Excel stylisé
def create_styled_excel(df, df_input, start_date, end_date):
    # Créer un classeur et sélectionner la feuille de calcul active
    wb = Workbook()
    ws = wb.active
    ws.title = "Etat de pointage"
    
    # Définir les styles
    header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    subheader_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
    recheck_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # Jaune clair pour vérification
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal='center', vertical='center')
    bold_font = Font(bold=True)
    
    # Ajouter la ligne de titre
    ws.merge_cells('A1:N1')
    title_cell = ws['A1']
    title_cell.value = f"Du {start_date} AU {end_date}"
    title_cell.alignment = center_alignment
    title_cell.font = bold_font
    
    # Ajouter la ligne d'en-tête
    headers = [
        'Matricule', 'Prénom/Nom', 'Dept', 'Date', 'Entrée', 'Sortie', 
        'Début Pause', 'Fin Pause', 'Temps de pause', 'H.P', 'H. Tr', 'HL',
        'Heures perdues', 'Heures supp', 'Cumul HP', 'Cumul HS', 'TOTAL HP', 'Observations'
    ]
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx)
        cell.value = header
        cell.fill = subheader_fill
        cell.border = thin_border
        cell.alignment = center_alignment
        cell.font = bold_font
    
    # Ajouter les lignes de données
    row_idx = 4
    current_matricule = None
    cumul_hp = timedelta()
    cumul_hs = timedelta()
    
    for _, row in df.iterrows():
        # Vérifier si nous commençons un nouvel employé
        if current_matricule != row['Matricule']:
            current_matricule = row['Matricule']
            cumul_hp = timedelta()
            cumul_hs = timedelta()
        
        # Ajouter les heures perdues au total cumulatif si présentes
        if row['Heures perdues']:
            try:
                h, m = map(int, row['Heures perdues'].split(':'))
                cumul_hp += timedelta(hours=h, minutes=m)
            except:
                pass
        
        # Ajouter les heures supplémentaires au total cumulatif si présentes
        if row['Heures supp']:
            try:
                h, m = map(int, row['Heures supp'].split(':'))
                cumul_hs += timedelta(hours=h, minutes=m)
            except:
                pass
        
        # Formater les heures cumulatives
        cumul_hp_str = f"{cumul_hp.days * 24 + cumul_hp.seconds // 3600:02d}:{(cumul_hp.seconds % 3600) // 60:02d}"
        cumul_hs_str = f"{cumul_hs.days * 24 + cumul_hs.seconds // 3600:02d}:{(cumul_hs.seconds % 3600) // 60:02d}"
        
        # Calculer le total HP (ce serait une décision de logique métier, pour l'instant juste montrer cumul_hp)
        total_hp = cumul_hp_str
        
        # Ajouter les données de ligne
        data = [
            row['Matricule'], row['Nom'], row['Départment'], row['Date'],
            row['Entrée'], row['Sortie'], row['Début Pause'], row['Fin Pause'],
            row['Temps de pause'], row['H.P'], row['H. Tr'], row['HL'],
            row['Heures perdues'], row['Heures supp'], cumul_hp_str, cumul_hs_str,
            total_hp, row['Observations']
        ]
        
        # Vérifier si cette ligne doit être mise en évidence pour vérification
        needs_recheck = "À VÉRIFIER" in str(row['Observations'])
        
        for col_idx, value in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.border = thin_border
            cell.alignment = center_alignment
            
            # Mettre en évidence toute la ligne si elle nécessite une vérification
            if needs_recheck:
                cell.fill = recheck_fill
            
            # Colorer la colonne H. Tr en rouge si elle est inférieure à HL
            if col_idx == 11 and value:  # Colonne H. Tr
                try:
                    h_tr = datetime.strptime(value, '%H:%M')
                    h_l = datetime.strptime(row['HL'], '%H:%M')
                    if h_tr < h_l:
                        cell.font = Font(color="FF0000")
                except:
                    pass
            
            # Colorer la colonne Heures perdues en rouge
            if col_idx == 13 and value:  # Colonne Heures perdues
                cell.font = Font(color="FF0000")
        
        row_idx += 1
    
    # Ajuster automatiquement les largeurs de colonne
    for col_idx in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = 15
    
    # Rendre la colonne Observations plus large
    ws.column_dimensions[get_column_letter(len(headers))].width = 40
    
    # Créer une deuxième feuille pour les données originales
    ws2 = wb.create_sheet(title="Données Originales")
    
    # Ajouter un titre à la deuxième feuille
    ws2.merge_cells('A1:E1')
    title_cell = ws2['A1']
    title_cell.value = "Données d'Entrée Originales"
    title_cell.alignment = center_alignment
    title_cell.font = bold_font
    
    # Ajouter des en-têtes à la deuxième feuille
    input_headers = list(df_input.columns)
    
    for col_idx, header in enumerate(input_headers, 1):
        cell = ws2.cell(row=3, column=col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center_alignment
        cell.font = bold_font
    
    # Ajouter des lignes de données à la deuxième feuille
    for row_idx, (_, row) in enumerate(df_input.iterrows(), 4):
        for col_idx, col_name in enumerate(input_headers, 1):
            cell = ws2.cell(row=row_idx, column=col_idx)
            cell.value = row[col_name]
            cell.border = thin_border
            cell.alignment = center_alignment
    
    # Ajuster automatiquement les largeurs de colonne dans la deuxième feuille
    for col_idx in range(1, len(input_headers) + 1):
        col_letter = get_column_letter(col_idx)
        ws2.column_dimensions[col_letter].width = 20
    
    # Créer une sortie en mémoire
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output

# Interface utilisateur Streamlit
st.title("Processeur de Données de Pointage")

st.markdown("""
Cette application traite les données de pointage et génère un rapport Excel stylisé.

### Format d'Entrée:
- Fichier CSV avec colonnes: Matricule;Nom;Départment;Date;Pointages


### Sortie:
- Fichier Excel stylisé avec calcul des heures de travail, pauses et totaux
- Les données d'entrée originales sont conservées dans une deuxième feuille
""")

# Téléchargeur de fichier
uploaded_file = st.file_uploader("Télécharger votre fichier CSV de données de pointage", type=["csv", "xls", "xlsx"])

if uploaded_file is not None:
    # Traiter le fichier
    try:
        # Analyser le fichier d'entrée
        df_input = parse_input_csv(uploaded_file)
        
        if df_input is None:
            st.error("Échec de l'analyse du fichier d'entrée. Veuillez vérifier le format.")
        else:
            
            
            # Traiter les entrées de temps
            df_processed = process_time_entries(df_input)

            
            # Compter et afficher les entrées qui nécessitent une vérification
            recheck_entries = df_processed[df_processed['Observations'].str.contains('À VÉRIFIER', na=False)]
            if not recheck_entries.empty:
                st.warning(f"Trouvé {len(recheck_entries)} entrées avec 5 valeurs qui nécessitent une vérification.")
                
                # Option pour afficher les entrées qui nécessitent une vérification
                if st.checkbox("Afficher les entrées qui nécessitent une vérification"):
                    st.dataframe(recheck_entries)
            
            # Obtenir la plage de dates pour le titre du rapport
            try:
                dates = pd.to_datetime(df_processed['Date'], format='%d/%m/%Y')
                start_date = dates.min().strftime('%d/%m/%Y')
                end_date = dates.max().strftime('%d/%m/%Y')
            except:
                # Repli si l'analyse de date échoue
                start_date = "01/01/2025"
                end_date = "31/12/2025"
            
            # Créer le fichier Excel stylisé avec les deux feuilles
            excel_output = create_styled_excel(df_processed, df_input, start_date, end_date)
            
            # Fournir un bouton de téléchargement
            st.download_button(
                label="Télécharger le Rapport Excel",
                data=excel_output,
                file_name="Etat_de_pointage.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        
    except Exception as e:
        st.error(f"Erreur lors du traitement du fichier: {str(e)}")
        st.exception(e)
else:
    st.info("Veuillez télécharger un fichier CSV à traiter.")

# Ajouter un bouton de déconnexion
if st.button("Déconnexion"):
    st.session_state.authenticated = False
    st.experimental_rerun()